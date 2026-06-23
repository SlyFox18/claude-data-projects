"""
Service Time Sheet Template Generator
Generates one .xlsm file per location, with one tab per technician.
Phase 1: openpyxl builds the workbook structure and styles.
Phase 2: Excel COM automation injects the VBA "Save As" macro button.
"""

import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
TECH_CSV    = os.path.join(BASE_DIR, "config", "TECH LEVELS.csv")
PAYROLL_CSV = os.path.join(BASE_DIR, "config", "Payroll Dates.csv")
OUTPUT_DIR  = os.path.join(BASE_DIR, "templates")

# ── Styles ─────────────────────────────────────────────────────────────────────
NAVY       = PatternFill("solid", fgColor="1F3864")
LIGHT_BLUE = PatternFill("solid", fgColor="DEEAF1")
YELLOW     = PatternFill("solid", fgColor="FFFF99")
ORANGE     = PatternFill("solid", fgColor="FFE0B2")
GREEN      = PatternFill("solid", fgColor="C6EFCE")
WHITE      = PatternFill("solid", fgColor="FFFFFF")
THIN       = Side(style="thin",   color="AAAAAA")
MED        = Side(style="medium", color="000000")
BOX        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center")
LOCKED     = Protection(locked=True)
UNLOCKED   = Protection(locked=False)

DATA_START = 8
DATA_END   = 41

VBA_CODE = """\
Sub SaveTimeSheet()
    Dim ws As Worksheet
    Set ws = ActiveSheet

    ' Require pay period before saving
    If Trim(CStr(ws.Range("M3").Value)) = "" Then
        MsgBox "Please select a Pay Period first." & Chr(13) & Chr(13) & _
               "Click the yellow Pay Period dropdown in cell M3.", _
               vbExclamation, "Select Pay Period First"
        ws.Range("M3").Select
        Exit Sub
    End If

    Dim suggestedName As String
    suggestedName = Trim(CStr(ws.Range("B56").Value))

    ' Open Save As dialog with the correct filename pre-filled
    Dim savePath As Variant
    savePath = Application.GetSaveAsFilename( _
        InitialFileName:=suggestedName & ".xlsx", _
        FileFilter:="Excel Workbook (*.xlsx),*.xlsx", _
        Title:="Save Time Sheet")

    If savePath = False Then Exit Sub

    ' Unprotect before saving so the submitted file arrives unprotected
    ' (recipients need full sort/filter access; the ribbon Sort & Filter button
    '  is greyed out on protected sheets regardless of AllowSorting/AllowFiltering)
    ws.Unprotect Password:=""

    ' Suppress the "save as xlsx loses macros" prompt
    Dim prevAlerts As Boolean
    prevAlerts = Application.DisplayAlerts
    Application.DisplayAlerts = False
    ActiveWorkbook.SaveAs Filename:=CStr(savePath), FileFormat:=51
    Application.DisplayAlerts = prevAlerts

    ' Re-protect the template so the tech's copy stays locked for data entry
    ws.Protect Password:="", DrawingObjects:=True, Contents:=True, Scenarios:=True, _
               AllowSorting:=True, AllowFiltering:=True, AllowInsertingRows:=True

    MsgBox "Time sheet saved!" & Chr(13) & CStr(savePath), _
           vbInformation, "Saved"
End Sub

Sub InsertDataRow()
    Dim ws As Worksheet
    Set ws = ActiveSheet

    ' Find the TOTAL HOURS row by scanning column A from row 8 down
    Dim totalsRow As Long
    totalsRow = 0
    Dim r As Long
    For r = 8 To 300
        If InStr(1, UCase(CStr(ws.Cells(r, 1).Value)), "TOTAL") > 0 Then
            totalsRow = r
            Exit For
        End If
    Next r

    If totalsRow = 0 Then
        MsgBox "Could not find the TOTAL HOURS row.", vbExclamation, "Error"
        Exit Sub
    End If

    ' Insert after the active row when inside the data area.
    ' Otherwise insert at totalsRow-1 (the spacer row position) so the spacer
    ' stays between the new row and TOTAL HOURS as the visual separator.
    Dim insertAt As Long
    Dim activeRow As Long
    activeRow = ActiveCell.Row
    If activeRow >= 8 And activeRow < totalsRow - 1 Then
        insertAt = activeRow + 1
    Else
        insertAt = totalsRow - 1
    End If

    ws.Unprotect Password:=""
    ws.Rows(insertAt).Insert Shift:=xlDown, CopyOrigin:=xlFormatFromAbove

    ' CopyOrigin copies cell formats but not row height — match the first data row
    ws.Rows(insertAt).RowHeight = ws.Rows(8).RowHeight

    ' Apply correct alternating fill: even rows = light blue, odd rows = white.
    ' CopyOrigin copies the same color as the row above instead of alternating,
    ' so we set it explicitly. Yellow input cells keep their yellow from CopyOrigin.
    Dim baseFill As Long
    If insertAt Mod 2 = 0 Then
        baseFill = RGB(222, 234, 241)   ' #DEEAF1 light blue
    Else
        baseFill = RGB(255, 255, 255)   ' white
    End If

    Dim cell As Range
    Dim c As Long
    For c = 1 To ws.UsedRange.Columns.Count
        Set cell = ws.Cells(insertAt, c)

        ' Fill: yellow input cells keep yellow, others get alternating color
        If cell.Interior.Color <> RGB(255, 255, 153) Then
            cell.Interior.Color = baseFill
        End If

        ' Thin border on all four sides matching template style (#AAAAAA)
        With cell.Borders
            .Item(xlEdgeLeft).LineStyle   = xlContinuous
            .Item(xlEdgeLeft).Weight      = xlThin
            .Item(xlEdgeLeft).Color       = RGB(170, 170, 170)
            .Item(xlEdgeTop).LineStyle    = xlContinuous
            .Item(xlEdgeTop).Weight       = xlThin
            .Item(xlEdgeTop).Color        = RGB(170, 170, 170)
            .Item(xlEdgeRight).LineStyle  = xlContinuous
            .Item(xlEdgeRight).Weight     = xlThin
            .Item(xlEdgeRight).Color      = RGB(170, 170, 170)
            .Item(xlEdgeBottom).LineStyle = xlContinuous
            .Item(xlEdgeBottom).Weight    = xlThin
            .Item(xlEdgeBottom).Color     = RGB(170, 170, 170)
        End With
    Next c

    ' ── Refresh TOTAL HOURS formulas to cover the newly inserted row ──────────
    ' totalsRow was found before the insert; the insert shifted it down by 1.
    ' newLastData is the row just above the spacer (two rows above totals).
    Dim newTotalsRow As Long
    Dim newLastData   As Long
    Dim sumCol        As Long
    Dim sumCell       As Range
    Dim colLetter     As String

    newTotalsRow = totalsRow + 1
    newLastData  = newTotalsRow - 2

    For sumCol = 1 To ws.UsedRange.Columns.Count
        Set sumCell = ws.Cells(newTotalsRow, sumCol)
        If sumCell.HasFormula Then
            colLetter = Mid(Split(ws.Columns(sumCol).Address(False, True), ":")(0), 2)
            sumCell.Formula = "=SUM(" & colLetter & "8:" & colLetter & newLastData & ")"
        End If
    Next sumCol

    ws.Protect Password:="", DrawingObjects:=True, Contents:=True, Scenarios:=True, _
               AllowSorting:=True, AllowFiltering:=True, AllowInsertingRows:=True

    ws.Cells(insertAt, 1).Select
End Sub
"""


# ── CSV readers ────────────────────────────────────────────────────────────────
def read_payroll_dates():
    periods = []
    with open(PAYROLL_CSV, newline="") as f:
        for row in csv.DictReader(f):
            start = row["Pay Period Start"].strip()
            end   = row["Pay Period End"].strip()
            periods.append({
                "label": f"{start} - {end}",
                "start": datetime.strptime(start, "%m/%d/%Y"),
                "end":   datetime.strptime(end,   "%m/%d/%Y"),
                "pay":   datetime.strptime(row["Pay Day"].strip(), "%m/%d/%Y"),
            })
    return periods


def read_techs_by_location():
    locations = {}
    with open(TECH_CSV, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            loc  = row["LOCATION"].strip()
            name = row["TECHNICIAN"].strip()
            if not loc or not name:
                continue
            locations.setdefault(loc, []).append({
                "num":        row.get("TECH #", "").strip(),
                "name":       name,
                "level":      row["LEVEL"].strip(),
                "mileage":    row.get("MILEAGE",    "No").strip().upper() == "YES",
                "shop_field": row.get("SHOP_FIELD", "No").strip().upper() == "YES",
            })
    return locations


# ── Column layout ──────────────────────────────────────────────────────────────
def get_col_config(has_mileage, has_shop_field):
    headers = [
        "CUSTOMER NAME", "MODEL", "RO#", "% WORK\nCOMPLETE",
        "*DRAW#1\nDATE", "*DRAW#1\nHRS", "*DRAW#2\nDATE", "*DRAW#2\nHRS",
        "*DRAW#3\nDATE", "*DRAW#3\nHRS", "*FINAL DRAW\nDATE", "*FINAL DRAW\nHRS",
    ]
    widths = [30, 14, 12, 12, 13, 10, 13, 10, 13, 10, 14, 12]
    cols = {"hrs": None, "shop": None, "field": None, "after": None, "mileage": None}

    if has_shop_field:
        headers += ["SHOP\nHRS", "FIELD\nHRS"]
        widths  += [12, 12]
        cols["shop"]  = 13
        cols["field"] = 14
        cols["after"] = 15
    else:
        headers += ["CURRENT DRAW/\nSHOP/FIELD HRS"]
        widths  += [18]
        cols["hrs"]   = 13
        cols["after"] = 14

    headers += ["AFTER\nHOURS"]
    widths  += [12]

    if has_mileage:
        headers += ["MILEAGE\n(x $1.85)"]
        widths  += [14]
        cols["mileage"] = cols["after"] + 1

    return headers, widths, cols


# ── Hidden support sheets ──────────────────────────────────────────────────────
def add_payroll_sheet(wb, periods):
    ws = wb.create_sheet("_PayrollDates")
    ws.sheet_state = "hidden"
    ws.append(["Period", "Start Date", "End Date", "Pay Day"])
    for p in periods:
        row = ws.max_row + 1
        ws.cell(row, 1, p["label"])
        for col, key in ((2, "start"), (3, "end"), (4, "pay")):
            c = ws.cell(row, col, p[key])
            c.number_format = "MM/DD/YYYY"


def add_lookup_sheet(wb, techs):
    ws = wb.create_sheet("_TechLookup")
    ws.sheet_state = "hidden"
    ws.append(["Tech #", "Name", "Level", "Mileage", "Shop/Field"])
    for t in techs:
        ws.append([
            t["num"], t["name"], t["level"],
            "Yes" if t["mileage"] else "No",
            "Yes" if t["shop_field"] else "No",
        ])


# ── Tech tab ───────────────────────────────────────────────────────────────────
def add_tech_sheet(wb, tech, location, num_periods):
    tab_name = tech["name"][:31].replace("/", "-")
    ws = wb.create_sheet(tab_name)

    has_mileage    = tech["mileage"]
    has_shop_field = tech["shop_field"]
    headers, widths, cols = get_col_config(has_mileage, has_shop_field)
    total_cols = len(headers)

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    yellow_input_cols = {v for v in cols.values() if v is not None}

    # ── Header block ──────────────────────────────────────────────────────────
    def hdr_label(cell, text):
        ws[cell].value      = text
        ws[cell].font       = Font(bold=True)
        ws[cell].protection = LOCKED

    def hdr_value(cell, val):
        ws[cell].value      = val
        ws[cell].font       = Font(bold=True)
        ws[cell].protection = LOCKED

    hdr_label("A2", "TECHNICIAN NAME:")
    hdr_value("B2", tech["name"])
    hdr_label("K2", "LOCATION:")
    hdr_value("M2", location.upper())

    hdr_label("A3", "TECH #:")
    hdr_value("B3", tech["num"] if tech["num"] else "")
    hdr_label("I3", "PAY PERIOD:")
    c = ws["M3"]
    c.fill, c.font      = YELLOW, Font(bold=True, color="000080")
    c.alignment, c.border = CENTER, BOX
    c.protection        = UNLOCKED

    hdr_label("A4", "TECH LEVEL:")
    hdr_value("B4", tech["level"])
    hdr_label("I4", "PAYROLL START DATE (MON):")
    c = ws["M4"]
    c.value         = '=IFERROR(INDEX(_PayrollDates!$B:$B,MATCH(M3,_PayrollDates!$A:$A,0)),"")'
    c.number_format = "MM/DD/YYYY"
    c.font          = Font(bold=True)
    c.protection    = LOCKED

    hdr_label("I5", "PAYROLL END DATE (SUN):")
    c = ws["M5"]
    c.value         = '=IFERROR(INDEX(_PayrollDates!$C:$C,MATCH(M3,_PayrollDates!$A:$A,0)),"")'
    c.number_format = "MM/DD/YYYY"
    c.font          = Font(bold=True)
    c.protection    = LOCKED

    ws.row_dimensions[6].height = 6

    # ── Column headers (row 7) ────────────────────────────────────────────────
    ws.row_dimensions[7].height = 36
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(7, col, hdr)
        c.fill       = NAVY
        c.font       = Font(color="FFFFFF", bold=True, size=9)
        c.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border     = BOX
        c.protection = LOCKED

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Explicit formats prevent Excel from drifting date-adjacent cells to Date type.
    DATE_ENTRY_COLS   = {5, 7, 9, 11}   # Draw1-3 date + Final Draw date
    NUMBER_ENTRY_COLS = {6, 8, 10, 12}  # Draw1-3 hrs + Final Draw hrs

    for row in range(DATA_START, DATA_END + 1):
        base_fill = LIGHT_BLUE if row % 2 == 0 else WHITE
        for col in range(1, total_cols + 1):
            c = ws.cell(row, col)
            c.border     = BOX
            c.protection = UNLOCKED
            if col in yellow_input_cols:
                c.fill, c.alignment = YELLOW, CENTER
                c.number_format = "0.00"
            elif col in DATE_ENTRY_COLS:
                c.fill, c.alignment = base_fill, LEFT
                c.number_format = "MM/DD/YYYY"
            elif col in NUMBER_ENTRY_COLS:
                c.fill, c.alignment = base_fill, LEFT
                c.number_format = "0.00"
            else:
                c.fill, c.alignment = base_fill, LEFT

    pct_dv = DataValidation(
        type="list", formula1='"0%,25%,50%,75%,100%"',
        allow_blank=True, showDropDown=False,
        error="Choose from the list", errorTitle="Invalid value",
        showErrorMessage=True,
    )
    pct_dv.sqref = f"D{DATA_START}:D{DATA_END}"
    ws.add_data_validation(pct_dv)

    period_dv = DataValidation(
        type="list",
        formula1=f"_PayrollDates!$A$2:$A${num_periods + 1}",
        allow_blank=True, showDropDown=False,
        error="Choose a pay period from the list", errorTitle="Invalid period",
        showErrorMessage=True,
    )
    period_dv.sqref = "M3"
    ws.add_data_validation(period_dv)

    # ── Totals row (43) ───────────────────────────────────────────────────────
    ws.row_dimensions[42].height = 6
    c = ws["A43"]
    c.value, c.font, c.protection = "TOTAL HOURS", Font(bold=True, size=11), LOCKED

    def make_total(col_idx):
        letter = ws.cell(1, col_idx).column_letter
        c = ws.cell(43, col_idx)
        c.value         = f"=SUM({letter}{DATA_START}:{letter}{DATA_END})"
        c.font          = Font(bold=True, size=11)
        c.fill          = GREEN
        c.alignment     = CENTER
        c.border        = BOX
        c.number_format = "0.00"
        c.protection    = LOCKED

    for key in ("hrs", "shop", "field", "after", "mileage"):
        if cols[key]:
            make_total(cols[key])

    # ── Holiday / PTO / No Pay ────────────────────────────────────────────────
    for row, lbl in ((44, "HOLIDAY HOURS"), (45, "PTO HOURS"),
                     (46, "NO PAY HOURS"), (47, "CIRCUMSTANTIAL HOURS")):
        c = ws.cell(row, 1, lbl)
        c.font, c.protection = Font(bold=True), LOCKED
        c = ws.cell(row, 2, 0)
        c.fill, c.border    = YELLOW, BOX
        c.alignment         = CENTER
        c.number_format     = "0.00"
        c.protection        = UNLOCKED

    c = ws["A48"]
    c.value      = "*FOR RECORD KEEPING PURPOSES ONLY"
    c.font       = Font(italic=True, size=9, color="888888")
    c.protection = LOCKED

    # ── Signature lines ───────────────────────────────────────────────────────
    ws.row_dimensions[50].height = 8
    sig_font = Font(bold=True, size=10)

    for row, lbl in ((51, "Employee Signature:"), (53, "Manager Signature:")):
        c = ws.cell(row, 1, lbl)
        c.font, c.protection = sig_font, LOCKED
        for col in range(2, 9):
            ws.cell(row, col).border     = Border(bottom=MED)
            ws.cell(row, col).protection = LOCKED
        c = ws.cell(row, 9, "Date:")
        c.font, c.protection = sig_font, LOCKED
        for col in range(10, 13):
            ws.cell(row, col).border     = Border(bottom=MED)
            ws.cell(row, col).protection = LOCKED
        ws.row_dimensions[row].height = 24

    ws.row_dimensions[52].height = 10

    # ── Save-as area ──────────────────────────────────────────────────────────
    ws.row_dimensions[55].height = 6
    loc_upper = location.upper()

    c = ws["A56"]
    c.value      = "SUGGESTED FILENAME:"
    c.font       = Font(bold=True, color="FF0000", size=10)
    c.protection = LOCKED

    # Formula — shows placeholder if no pay period selected yet
    c = ws["B56"]
    c.value = (
        f'=IF(M3="","<- Select Pay Period in M3 first",'
        f'"{loc_upper} TECHS "&TEXT(M4,"MM.DD.YY")&" to "&TEXT(M5,"MM.DD.YY"))'
    )
    c.font       = Font(bold=True, color="FF0000", size=11)
    c.fill       = PatternFill("solid", fgColor="FFF2CC")
    c.protection = UNLOCKED   # unlocked so it can be copied as fallback
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 44)

    # Row 57 is reserved for the Save button (added via COM in Phase 2)
    ws.row_dimensions[57].height = 30

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left   = 0.25
    ws.page_margins.right  = 0.25
    ws.page_margins.top    = 0.25
    ws.page_margins.bottom = 0.25
    last_col = ws.cell(1, total_cols).column_letter
    ws.print_area = f"A1:{last_col}56"

    # NOTE: sheet protection is applied in Phase 2 (after button is added)


# ── Phase 2: VBA macro + button via Excel COM ──────────────────────────────────
def add_vba_macros(output_dir, locations):
    try:
        import win32com.client as win32
        import pythoncom
    except ImportError:
        print("\n  WARNING: pywin32 not installed — VBA button skipped.")
        print("  Run: pip install pywin32")
        return

    print("\nPhase 2 - Adding VBA Save button (this opens Excel in the background)...")

    pythoncom.CoInitialize()
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible        = False
    excel.DisplayAlerts  = False

    try:
        for loc in sorted(locations):
            xlsx_path = os.path.abspath(
                os.path.join(output_dir, f"{loc.upper()} - Service Time Sheet Template.xlsx")
            )
            xlsm_path = xlsx_path.replace(".xlsx", ".xlsm")
            print(f"  {loc}...", end="", flush=True)

            wb = excel.Workbooks.Open(xlsx_path)

            # Add the VBA module
            try:
                mod = wb.VBProject.VBComponents.Add(1)   # 1 = vbext_ct_StdModule
                mod.Name = "SvcTimeSheetMacros"
                mod.CodeModule.AddFromString(VBA_CODE)
            except Exception as e:
                print(f" SKIPPED - VBA access blocked ({e})")
                print("    Fix: Excel -> File -> Options -> Trust Center -> Macro Settings")
                print("         -> enable 'Trust access to the VBA project object model'")
                wb.Close(False)
                continue

            # Add button + protection to each tech sheet
            for i in range(1, wb.Worksheets.Count + 1):
                ws = wb.Worksheets(i)
                if ws.Name.startswith("_"):
                    continue

                # Place the Save button in row 57
                btn_left = ws.Range("A57").Left
                btn_top  = ws.Range("A57").Top + 2
                btn = ws.Buttons().Add(btn_left, btn_top, 220, 26)
                btn.OnAction = "SaveTimeSheet"
                btn.Caption  = "Save As Correct Name..."
                btn.Font.Bold = True
                btn.Font.Size = 10

                # Insert Row button — placed immediately to the right of Save button
                ins_btn = ws.Buttons().Add(btn_left + 230, btn_top, 160, 26)
                ins_btn.OnAction = "InsertDataRow"
                ins_btn.Caption  = "Insert Row Below"
                ins_btn.Font.Bold = True
                ins_btn.Font.Size = 10

                # Enable AutoFilter on header row so filter arrows survive protection
                ws.Range("A7").AutoFilter()

                # Protect sheet — locked cells protected, unlocked cells editable
                # AllowSorting + AllowFiltering let managers sort/filter the RO rows
                # AllowInsertingRows lets managers add rows when a tech has more than 34 ROs
                ws.Protect(
                    Password="",
                    DrawingObjects=True,
                    Contents=True,
                    Scenarios=True,
                    AllowSorting=True,
                    AllowFiltering=True,
                    AllowInsertingRows=True,
                )

            # Save as macro-enabled workbook
            wb.SaveAs(xlsm_path, 52)   # 52 = xlOpenXMLWorkbookMacroEnabled
            wb.Close(False)
            os.remove(xlsx_path)
            print(f" OK -> {os.path.basename(xlsm_path)}")

    finally:
        excel.Quit()
        pythoncom.CoUninitialize()

    print("Phase 2 complete.\n")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print("Reading source files...")
    periods   = read_payroll_dates()
    locations = read_techs_by_location()
    print(f"  {len(periods)} pay periods loaded")
    print(f"  {sum(len(v) for v in locations.values())} techs across {len(locations)} locations\n")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Phase 1 — build workbook structure with openpyxl
    print("Phase 1 — Building workbook structure...")
    for loc in sorted(locations):
        techs = locations[loc]
        wb = Workbook()
        wb.remove(wb.active)

        for tech in techs:
            add_tech_sheet(wb, tech, loc, len(periods))

        add_payroll_sheet(wb, periods)
        add_lookup_sheet(wb, techs)

        filename = f"{loc.upper()} - Service Time Sheet Template.xlsx"
        wb.save(os.path.join(OUTPUT_DIR, filename))
        print(f"  {filename}  ({len(techs)} tabs)")

    # Phase 2 — inject VBA macro + Save button via Excel COM
    add_vba_macros(OUTPUT_DIR, locations)

    print(f"Done. Templates saved to:\n  {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
