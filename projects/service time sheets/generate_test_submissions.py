"""
Generates realistic test time sheet submission files for Fabric Dataflow development.
Mimics what a service manager submits after filling out the .xlsm template.

Run once, then upload the output files to their location folders in SharePoint
(not _Template — the actual location folders) to test the dataflow.
"""

import os
import random
from datetime import date

from openpyxl import Workbook

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "test-submissions")

# Pay period — matches the one in the test email (04/20/2026 - 05/03/2026)
PAY_PERIOD_LABEL = "04/20/2026 - 05/03/2026"
PAY_START        = date(2026, 4, 20)
PAY_END          = date(2026, 5, 3)

CUSTOMERS = [
    "H&H Farms", "Red River Ag", "Bravo Cotton", "Mesa Verde Ranch",
    "Double T Farms", "Lone Star Grain", "Garcia Family Farm", "Plains Cotton Coop",
    "Rio Grande Farms", "High Mesa Agriculture", "Sunset Irrigation", "Panhandle Farms",
    "Triple C Ranch", "Yellowhouse Canyon Ag", "Blanco River Farms",
]

MODELS = [
    "9R 590", "8R 310", "7R 290", "6M 215", "6120M", "5115M",
    "S790", "8400R", "9620RX", "6155M", "5075E", "7230R",
    "9RX 640", "8R 250", "6130R",
]

# ── Column config (mirrors generate_templates.py exactly) ─────────────────────
def get_col_config(has_mileage, has_shop_field):
    headers = [
        "CUSTOMER NAME", "MODEL", "RO#", "% WORK\nCOMPLETE",
        "*DRAW#1\nDATE", "*DRAW#1\nHRS", "*DRAW#2\nDATE", "*DRAW#2\nHRS",
        "*DRAW#3\nDATE", "*DRAW#3\nHRS", "*FINAL DRAW\nDATE", "*FINAL DRAW\nHRS",
    ]
    cols = {"hrs": None, "shop": None, "field": None, "after": None, "mileage": None}

    if has_shop_field:
        headers += ["SHOP\nHRS", "FIELD\nHRS"]
        cols["shop"]  = 13
        cols["field"] = 14
        cols["after"] = 15
    else:
        headers += ["CURRENT DRAW/\nSHOP/FIELD HRS"]
        cols["hrs"]   = 13
        cols["after"] = 14

    headers += ["AFTER\nHOURS"]

    if has_mileage:
        headers += ["MILEAGE\n(x $1.85)"]
        cols["mileage"] = cols["after"] + 1

    return headers, cols


# ── Test locations — covers the three column layout variants ──────────────────
# Lorenzo:   no mileage, no shop/field  (most locations)
# Brownfield: mileage, no shop/field   (tests mileage column)
# Levelland:  mileage, no shop/field   (second mileage location for volume)
LOCATIONS = {
    "Lorenzo": {
        "mileage":    False,
        "shop_field": False,
        "techs": [
            {"name": "Nick Nunley",      "num": "T9105", "level": "5"},
            {"name": "Kevin Hendley",    "num": "T9107", "level": "5"},
            {"name": "Larry Davis",      "num": "T9108", "level": "5"},
            {"name": "Andrew Martin",    "num": "T9112", "level": "4"},
            {"name": "Angel Reyes",      "num": "T9121", "level": "4"},
            {"name": "Johnnie Fetsch",   "num": "T9129", "level": "5"},
        ],
    },
    "Brownfield": {
        "mileage":    True,
        "shop_field": False,
        "techs": [
            {"name": "Josh Jeffcoat",    "num": "T1030", "level": "5"},
            {"name": "Antonio Guzman",   "num": "T1506", "level": "4"},
            {"name": "Dylan Wallin",     "num": "T1143", "level": "3"},
            {"name": "Garrett Muncy",    "num": "T1130", "level": "4"},
            {"name": "Nico Lee",         "num": "T1512", "level": "4"},
        ],
    },
    "Levelland": {
        "mileage":    True,
        "shop_field": False,
        "techs": [
            {"name": "Ricky Dudley",     "num": "T1225", "level": "5"},
            {"name": "Clayton Madewell", "num": "T1134", "level": "5"},
            {"name": "George Patterson", "num": "T1514", "level": "4"},
            {"name": "Justin Smith",     "num": "",      "level": "3"},
            {"name": "Payton Oliver",    "num": "",      "level": "3"},
        ],
    },
}


# ── Helpers ───────────────────────────────────────────────────────────────────
def rand_ro():
    return random.randint(10000, 99999)

def rand_hrs(lo=1.0, hi=8.0):
    # Round to nearest 0.5
    return round(random.uniform(lo, hi) * 2) / 2

def rand_pct():
    return random.choice(["25%", "50%", "75%", "100%"])


# ── Build one tech tab ────────────────────────────────────────────────────────
def build_tech_sheet(ws, tech, location, headers, cols, rng):
    # Header block — mirrors the template's cell layout exactly
    ws["B2"] = tech["name"]
    ws["B3"] = tech["num"]
    ws["B4"] = tech["level"]
    ws["M2"] = location.upper()
    ws["M3"] = PAY_PERIOD_LABEL
    c = ws["M4"]; c.value = PAY_START; c.number_format = "MM/DD/YYYY"
    c = ws["M5"]; c.value = PAY_END;   c.number_format = "MM/DD/YYYY"

    # Column headers at row 7
    for col, hdr in enumerate(headers, 1):
        ws.cell(7, col, hdr)

    # RO data rows (rows 8–41)
    num_ros    = rng.randint(4, 9)
    total_hrs  = 0.0
    total_after = 0.0

    for i in range(num_ros):
        row = 8 + i

        cur_hrs   = rand_hrs(2.0, 14.0)
        after_hrs = rng.choice([0.0, 0.0, 0.0, rand_hrs(0.5, 2.0)])
        draw1_hrs = rand_hrs(0.5, min(cur_hrs, 4.0))

        ws.cell(row, 1, rng.choice(CUSTOMERS))
        ws.cell(row, 2, rng.choice(MODELS))
        ws.cell(row, 3, rand_ro())
        ws.cell(row, 4, rand_pct())
        c = ws.cell(row, 5, PAY_START); c.number_format = "MM/DD/YYYY"
        ws.cell(row, 6, draw1_hrs)

        # Second draw on roughly half the ROs
        if rng.random() > 0.5:
            c = ws.cell(row, 7, PAY_END); c.number_format = "MM/DD/YYYY"
            ws.cell(row, 8, rand_hrs(0.5, 2.0))

        if cols["hrs"] is not None:
            ws.cell(row, cols["hrs"], cur_hrs)
        elif cols["shop"] is not None:
            shop  = round(cur_hrs * rng.uniform(0.5, 0.8), 1)
            field = round(cur_hrs - shop, 1)
            ws.cell(row, cols["shop"],  shop)
            ws.cell(row, cols["field"], field)

        ws.cell(row, cols["after"], after_hrs)

        if cols["mileage"] is not None:
            miles = rng.choice([0, 0, rng.randint(15, 120)])
            ws.cell(row, cols["mileage"], miles)

        total_hrs   += cur_hrs
        total_after += after_hrs

    # Totals row (43)
    ws.cell(43, 1, "TOTAL HOURS")
    if cols["hrs"] is not None:
        ws.cell(43, cols["hrs"], round(total_hrs, 2))
    ws.cell(43, cols["after"], round(total_after, 2))

    # Holiday / PTO / No Pay / Circumstantial (rows 44–47)
    for row, lbl in (
        (44, "HOLIDAY HOURS"),
        (45, "PTO HOURS"),
        (46, "NO PAY HOURS"),
        (47, "CIRCUMSTANTIAL HOURS"),
    ):
        ws.cell(row, 1, lbl)
        ws.cell(row, 2, 0)


# ── Build one location workbook ───────────────────────────────────────────────
def create_submission(location, spec, seed=42):
    rng = random.Random(seed)   # fixed seed → reproducible, varied per location

    has_mileage    = spec["mileage"]
    has_shop_field = spec["shop_field"]
    headers, cols  = get_col_config(has_mileage, has_shop_field)

    wb = Workbook()
    wb.remove(wb.active)

    for tech in spec["techs"]:
        ws = wb.create_sheet(tech["name"])
        build_tech_sheet(ws, tech, location, headers, cols, rng)

    start_str = PAY_START.strftime("%m.%d.%y")
    end_str   = PAY_END.strftime("%m.%d.%y")
    filename  = f"{location.upper()} TECHS {start_str} to {end_str}.xlsx"

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    return filename


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Generating test submissions  (pay period: {PAY_PERIOD_LABEL})\n")
    for i, (location, spec) in enumerate(LOCATIONS.items()):
        filename = create_submission(location, spec, seed=42 + i)
        col_desc = []
        if spec["mileage"]:    col_desc.append("mileage")
        if spec["shop_field"]: col_desc.append("shop/field")
        layout = f"  [{', '.join(col_desc) or 'standard layout'}]"
        print(f"  {filename}  ({len(spec['techs'])} techs){layout}")

    print(f"\nSaved to: {OUTPUT_DIR}")
    print("\nNext steps:")
    print("  1. Upload each file to its matching location folder in SharePoint")
    print("     (Service Time Sheets/Lorenzo/, /Brownfield/, /Levelland/)")
    print("     NOT the _Template subfolder — the location root folder")
    print("  2. Build df_ServiceTimeSheets_Raw in Fabric and test against these files")


if __name__ == "__main__":
    main()
